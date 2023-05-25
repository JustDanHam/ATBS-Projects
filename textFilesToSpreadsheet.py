#! python3
# textFilesToSpreadsheet.py - reads in the contents of several text files
# and inserts those contents into a spreadsheet, with one line of text
# per row. The lines of the first text file will be in the cells of column
# A, the lines of the second text file will be in the cells of column B, etc.

import openpyxl
from openpyxl.styles import Font
import os

filenames = ['ColA.txt', 'ColB.txt']

wb = openpyxl.Workbook()
sheet = wb.active
sheet.freeze_panes = 'A2'
headerFont = Font(bold=True)

for col, filename in enumerate(filenames):
    colHeader = sheet.cell(column=col+1, row=1)
    colHeader.font = headerFont
    colHeader.value = os.path.splitext(filename)[0]

    textFile = open(filename)
    colData = textFile.read().splitlines()

    for row, cellData in enumerate(colData):
        sheet.cell(column=col+1, row=row+2).value = cellData
    textFile.close()

wb.save(f'{input("Save file as: ")}.xlsx')
print('Done. Spreadsheet created.')
