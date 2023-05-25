#! python3
# spreadsheetToTextFiles.py - takes a spreadsheet and writes the data in each
# column to a seperate text file.
# Each text file will be named after the value in row 1 of the column.


import openpyxl

filename = input('Choose spreadsheet: ')
wb = openpyxl.load_workbook(filename, data_only=True)
sheet = wb.active

for col in range(1, sheet.max_column + 1):
    colHeader = sheet.cell(column=col, row=1).value
    textFile = open(f'{filename[:-5]}{colHeader}', mode='a')
    for row in range(1, sheet.max_row +1):
        cellData = sheet.cell(column=col, row=row).value
        textFile.write(f'{cellData}\n')
    textFile.close()

print('Text files have been created')
