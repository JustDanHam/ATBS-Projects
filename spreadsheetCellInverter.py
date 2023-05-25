#! python3
# spreadsheetCellInverter.py - a program that inverts the rows and columns
# of a spreadsheet.

import openpyxl


filename = input('Insert directory of spreadsheet to invert: ')

wb = openpyxl.load_workbook(filename)
sheet = wb.active
invertedSheet = wb.create_sheet(title=f'Inverted {sheet.title}')

for ro in range(1, sheet.max_row + 1):
    for col in range(1, sheet.max_column + 1):
        invertedSheet.cell(row=col, column=ro).value = sheet.cell(row=ro, column=col).value
        print(f'Cell {ro},{col} inverted.')

wb.save(f'{filename[:-5]}Inverted.xlsx')
print(f'Inversion complete. Added as sheet "{invertedSheet.title}" and saved as "{filename[:-5]}Inverted.xlsx"')
