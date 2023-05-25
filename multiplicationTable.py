#! python3
# multiplicationTable.py - takes a number N from the command line and
# creates an NxN mulitplication table in an Excel spreadsheet.

import sys
import openpyxl
from openpyxl.styles import Font

N = int(sys.argv[1])

wb = openpyxl.Workbook()
sheet = wb.active
labelFont = Font(bold=True)

for i in range(1, N+2):
    for j in range(1, N+2):
        if i == 1:
            if j == N + 1:
                continue
            sheet.cell(row=i, column=j+1).font = labelFont
            sheet.cell(row=i, column=j+1).value = j
        if j == 1:
            if i == N + 1:
                continue
            sheet.cell(row=i+1, column=j).font = labelFont
            sheet.cell(row=i+1, column=j).value = i
        if i > 1 and j > 1:
            cell1 = sheet.cell(row=1, column=j)
            cell2 = sheet.cell(row=i, column=1)
            sheet.cell(row=i, column=j).value = f'=PRODUCT({cell1.coordinate},{cell2.coordinate})'

wb.save(f'multiplicationTable1to{str(N)}.xlsx')
print(f'Your multiplication table "multiplicationTable1to{str(N)}.xlsx" has been created in the cwd.')
