#! python3
# excelToCsvConverter.py - Converts all excel files in the current working
# directory to CSV files.

import os
import openpyxl
import csv


os.makedirs('ExcelToCSV', exist_ok=True)

for excelFile in os.listdir('.'):
    # Skip non-xlsx files, load the workbook object.
    if not excelFile.endswith('.xlsx'):
        continue
    wb = openpyxl.load_workbook(excelFile)

    for sheetName in wb.sheetnames:
        # Loop through every sheet in the workbook.
        sheet = wb[sheetName]

        # Create the CSV filename from the Excel filename and sheet title.
        csvFilename = os.path.join('ExcelToCSV', f'{excelFile[:-5]}-{sheetName}.csv')
        csvFileObj = open(csvFilename, 'w', newline='')

        # Create the csv.writer object for this CSV file.
        csvWriter = csv.writer(csvFileObj)

        # Loop through every row in the sheet.
        for rowNum in range(1, sheet.max_row + 1):
            rowData = []        # append each cell to this list
            # Loop through each cell in the row.
            for colNum in range(1, sheet.max_column + 1):
                # Append each cell's data to rowData.
                rowData.append(sheet.cell(column=colNum, row=rowNum).value)
            # Write the rowData list to the CSV file.
            csvWriter.writerow(rowData)

        csvFileObj.close()
